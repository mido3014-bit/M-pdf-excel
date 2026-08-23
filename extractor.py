# -*- coding: utf-8 -*-
"""
extractor.py
------------
استخراج جدول "Bill of Material of Structure" وجدول "Structure Quantity Required"
من ملفات PDF لرسومات إنشائية (شيتات حديد) وتصديرهم إلى ملف Excel بنفس الشكل.

الفكرة العامة:
1) نفتح ملف الـ PDF بواسطة pdfplumber.
2) نبحث في كل صفحة عن الجداول (باستخدام خطوط الجدول الفعلية Grid Lines لأن
   هذا النوع من رسومات الـ CAD يخرج بجداول محددة بخطوط واضحة).
3) نحدد الجدول الكبير (Bill of Material) عن طريق البحث عن كلمات مفتاحية
   في الصف الأول مثل: ERECTION MARK, MEMBER SIZE, QUANTITY, UNIT WT ...
4) نحدد الجدول الصغير (Structure Quantity Required) بنفس الطريقة.
5) ننظف البيانات ونصدرها إلى Excel بتنسيق قريب من شكل المصدر.

ملاحظة هامة:
رسومات الـ CAD المختلفة قد تخرج بصيغ PDF مختلفة قليلاً (بعضها نص حقيقي
قابل للنسخ، وبعضها صورة ممسوحة/Raster). هذا السكريبت مصمم للحالة الأولى
(PDF بها طبقة نصية حقيقية وخطوط جدول واضحة) وهي الحالة الأشيع في رسومات
الـ Tekla / AutoCAD / MicroStation المُصدَّرة كـ PDF مباشرة.
إذا كان الـ PDF عبارة عن صورةممسوحة (Scanned) سيحتاج الأمر خطوة OCR إضافية
(موضح فى الأسفل كيف تفعّلها).
"""

import re
import sys
from dataclasses import dataclass, field
from typing import List, Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# إعدادات
# ---------------------------------------------------------------------------

BOM_HEADER_KEYWORDS = [
    "erection", "member size", "approx", "quantity", "unit wt", "total wt", "per piece"
]

QTY_TABLE_KEYWORDS = [
    "structure quantity required", "structure wt"
]

TABLE_SETTINGS = {
    "vertical_strategy": "lines",
    "horizontal_strategy": "lines",
    "snap_tolerance": 4,
    "join_tolerance": 4,
    "intersection_tolerance": 4,
}

# فallback: لو الجدول مفيهوش خطوط واضحة كفاية، نجرب استراتيجية "text"
TABLE_SETTINGS_FALLBACK = {
    "vertical_strategy": "text",
    "horizontal_strategy": "text",
    "snap_tolerance": 5,
    "join_tolerance": 5,
    "text_tolerance": 3,
}


@dataclass
class ExtractedTable:
    page_number: int
    rows: List[List[str]] = field(default_factory=list)

    def clean(self):
        cleaned = []
        for row in self.rows:
            cleaned_row = [
                (cell or "").replace("\n", " ").strip() if cell is not None else ""
                for cell in row
            ]
            # تجاهل الصفوف الفاضية تماما
            if any(c for c in cleaned_row):
                cleaned.append(cleaned_row)
        self.rows = cleaned
        return self


def _row_matches_keywords(row: List[str], keywords: List[str]) -> int:
    """يرجع عدد الكلمات المفتاحية اللي اتلاقت في الصف (لتقييم مدى تطابق الجدول)."""
    joined = " ".join(row).lower()
    return sum(1 for kw in keywords if kw in joined)


def find_tables_in_pdf(pdf_path: str) -> List[ExtractedTable]:
    """يمر على كل صفحات الـ PDF ويرجع كل الجداول المكتشفة (طبقة نص فقط، بدون OCR)."""
    found: List[ExtractedTable] = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables(TABLE_SETTINGS)
            if not tables:
                tables = page.extract_tables(TABLE_SETTINGS_FALLBACK)
            for t in tables:
                if t and len(t) >= 1:
                    found.append(ExtractedTable(page_number=i, rows=t).clean())
    return found


def _looks_scanned_or_empty(tables: List[ExtractedTable]) -> bool:
    """
    تقييم بسيط: لو مفيش جداول خالص، أو الجداول المكتشفة فيها خلايا فاضية
    غالبًا (أكتر من 60%)، يبقى على الأغلب الملف صورة ممسوحة (Scanned)
    ومحتاج OCR بدل الاعتماد على طبقة النص.
    """
    if not tables:
        return True
    total_cells = 0
    empty_cells = 0
    for t in tables:
        for row in t.rows:
            for cell in row:
                total_cells += 1
                if not cell.strip():
                    empty_cells += 1
    if total_cells == 0:
        return True
    return (empty_cells / total_cells) > 0.6


def find_tables_with_ocr_fallback(pdf_path: str, force_ocr: bool = False) -> List[ExtractedTable]:
    """
    الدالة الموصى باستخدامها بدل find_tables_in_pdf مباشرة:
    1) تحاول الاستخراج العادي من طبقة النص أولًا (أسرع وأدق لو الملف مش صورة).
    2) لو النتيجة فاضية/ضعيفة (يبدو أن الملف Scanned) أو force_ocr=True،
       تلجأ تلقائيًا لاستخراج OCR (ocr_extractor.py) وترجع نتيجته بنفس الشكل.
    """
    tables = [] if force_ocr else find_tables_in_pdf(pdf_path)

    if force_ocr or _looks_scanned_or_empty(tables):
        try:
            from ocr_extractor import find_tables_via_ocr
        except ImportError as e:
            raise ImportError(
                "الملف يبدو أنه صورة ممسوحة (Scanned) ويحتاج OCR، لكن مكتبات الـ OCR غير "
                "مثبتة. ثبّت المتطلبات من requirements.txt وتأكد من تثبيت tesseract-ocr "
                "و poppler-utils على مستوى النظام."
            ) from e

        ocr_tables = find_tables_via_ocr(pdf_path)
        ocr_as_extracted = [
            ExtractedTable(page_number=t.page_number, rows=t.rows).clean() for t in ocr_tables
        ]
        # لو الاستخراج العادي رجع حاجة أقوى من الـ OCR (نادر) نفضلها، غير كده نستخدم OCR
        if len(ocr_as_extracted) >= len(tables):
            return ocr_as_extracted

    return tables


def pick_bom_table(tables: List[ExtractedTable]) -> Optional[ExtractedTable]:
    """يختار أفضل جدول يمثل Bill of Material بناء على تطابق الكلمات المفتاحية وعدد الصفوف."""
    best = None
    best_score = 0
    for t in tables:
        # نفحص أول 3 صفوف كهيدر محتمل
        header_zone = t.rows[:3]
        score = sum(_row_matches_keywords(r, BOM_HEADER_KEYWORDS) for r in header_zone)
        score += len(t.rows) * 0.1  # نفضل الجدول الأطول لو التطابق متقارب
        if score > best_score:
            best_score = score
            best = t
    return best if best_score >= 3 else best  # نرجع أفضل حاجة لقيناها حتى لو التطابق ضعيف


def pick_qty_table(tables: List[ExtractedTable]) -> Optional[ExtractedTable]:
    """يختار جدول (STRUCTURE QUANTITY REQUIRED / STRUCTURE WT)."""
    best = None
    best_score = 0
    for t in tables:
        joined_all = " ".join(" ".join(r) for r in t.rows).lower()
        score = sum(1 for kw in QTY_TABLE_KEYWORDS if kw in joined_all)
        if score > best_score:
            best_score = score
            best = t
    return best if best_score > 0 else None


def parse_qty_table(t: ExtractedTable) -> dict:
    """
    يحاول استخراج:
      - structure_mark  (مثال: K77 -4 NO'S)
      - quantity_required (مثال: 4 NO.S)
      - weight_formula   (مثال: 4 X 444.58 Kg)
      - total_weight     (مثال: 1778.32 Kg)
    من نص الجدول الصغير بشكل مرن (regex) بدل الاعتماد فقط على شكل الخلايا.
    """
    full_text = " | ".join(" ".join(r) for r in t.rows)

    result = {
        "structure_mark": None,
        "quantity_required": None,
        "weight_per_unit_formula": None,
        "total_weight": None,
        "raw_text": full_text,
    }

    # STRUCTURE QUANTITY REQUIRED   4 NO.S
    m = re.search(r"QUANTITY REQUIRED\D*([\d.,]+\s*NO\.?S?)", full_text, re.IGNORECASE)
    if m:
        result["quantity_required"] = m.group(1).strip()

    # STRUCTURE WT = 4 X 444.58 Kg  1778.32 Kg
    m = re.search(
        r"STRUCTURE\s*WT\D*([\d.,]+)\s*X\s*([\d.,]+)\s*Kg.*?([\d.,]+)\s*Kg",
        full_text,
        re.IGNORECASE,
    )
    if m:
        qty, unit_wt, total = m.groups()
        result["weight_per_unit_formula"] = f"{qty} X {unit_wt} Kg"
        result["total_weight"] = total

    # الماركة نفسها (زي K77 -4 NO'S) عادة أول خلية في أول صف
    if t.rows:
        first_cell = t.rows[0][0] if t.rows[0] else ""
        if first_cell:
            result["structure_mark"] = first_cell.strip()

    return result


# ---------------------------------------------------------------------------
# التصدير إلى Excel
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
BOLD_FONT = Font(bold=True)


def _write_table(ws, start_row: int, rows: List[List[str]], bold_keywords=("TOTAL WEIGHT",)):
    """يكتب جدول (list of rows) في worksheet ابتداء من start_row، مع تنسيق بسيط."""
    r = start_row
    for row_idx, row in enumerate(rows):
        is_header = row_idx == 0
        is_total_row = any(kw.lower() in " ".join(row).lower() for kw in bold_keywords)
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=_maybe_number(value))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if is_header:
                cell.fill = HEADER_FILL
                cell.font = BOLD_FONT
            elif is_total_row:
                cell.fill = TOTAL_FILL
                cell.font = BOLD_FONT
        r += 1
    return r  # الصف التالي الفاضي


def _maybe_number(value: str):
    """يحاول تحويل النص لرقم لو ممكن، عشان الخلايا في إكسل تبقى أرقام حقيقية (قابلة لعمليات حسابية)."""
    if value is None:
        return value
    v = value.strip()
    if v == "":
        return v
    v_norm = v.replace(",", "")
    try:
        if "." in v_norm:
            return float(v_norm)
        return int(v_norm)
    except ValueError:
        return value


def export_to_excel(bom_table: Optional[ExtractedTable],
                     qty_info: Optional[dict],
                     output_path: str,
                     sheet_title: str = "Bill of Material"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] if sheet_title else "Bill of Material"

    row_cursor = 1

    ws.cell(row=row_cursor, column=1, value="BILL OF MATERIAL OF STRUCTURE").font = Font(bold=True, size=13)
    row_cursor += 2

    if bom_table and bom_table.rows:
        row_cursor = _write_table(ws, row_cursor, bom_table.rows)
    else:
        ws.cell(row=row_cursor, column=1, value="⚠ لم يتم العثور على جدول Bill of Material في هذا الملف")
        row_cursor += 1

    row_cursor += 2

    ws.cell(row=row_cursor, column=1, value="STRUCTURE QUANTITY").font = Font(bold=True, size=13)
    row_cursor += 2

    if qty_info:
        labels = [
            ("Structure Mark", qty_info.get("structure_mark")),
            ("Structure Quantity Required", qty_info.get("quantity_required")),
            ("Weight Formula", qty_info.get("weight_per_unit_formula")),
            ("Total Weight (Kg)", qty_info.get("total_weight")),
        ]
        for label, val in labels:
            c1 = ws.cell(row=row_cursor, column=1, value=label)
            c2 = ws.cell(row=row_cursor, column=2, value=_maybe_number(val) if val else "")
            c1.font = BOLD_FONT
            c1.border = THIN_BORDER
            c2.border = THIN_BORDER
            row_cursor += 1
    else:
        ws.cell(row=row_cursor, column=1, value="⚠ لم يتم العثور على جدول Structure Quantity في هذا الملف")

    # ضبط عرض الأعمدة تلقائيًا (تقريبي)
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 45)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# نقطة تشغيل من سطر الأوامر (CLI)
# ---------------------------------------------------------------------------

def process_pdf(pdf_path: str, output_xlsx: str, force_ocr: bool = False):
    tables = find_tables_with_ocr_fallback(pdf_path, force_ocr=force_ocr)
    bom = pick_bom_table(tables)
    qty_table = pick_qty_table(tables)
    qty_info = parse_qty_table(qty_table) if qty_table else None
    export_to_excel(bom, qty_info, output_xlsx)
    return bom, qty_info


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("الاستخدام: python extractor.py input.pdf output.xlsx [--force-ocr]")
        sys.exit(1)
    in_path, out_path = sys.argv[1], sys.argv[2]
    force = "--force-ocr" in sys.argv[3:]
    bom, qty_info = process_pdf(in_path, out_path, force_ocr=force)
    print(f"تم الحفظ في: {out_path}")
    print(f"عدد صفوف جدول BOM المستخرج: {len(bom.rows) if bom else 0}")
    print(f"بيانات جدول الكمية: {qty_info}")
